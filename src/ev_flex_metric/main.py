import json
import math
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Tuple

import avro.schema
from avro.datafile import DataFileWriter
from avro.io import DatumWriter
import pytz
import openpyxl

from ev_flex_metric.ranges import IntRangeInBlock, DecimalRangeInBlock, DecimalInstantInBlock, IntInstantInBlock


def times_ranges_overlap(start1: datetime, end1: datetime, start2: datetime, end2: datetime):
    return not(end1 <= start2 or start1 >= end2)


@dataclass
class BlockMetadata:
    """Metadata for a profile in time.

    The block has a start time, an end time and a step duration splitting the start
    and end time in equal-sized blocks.
    """
    start_time: datetime
    end_time: datetime
    step_duration: timedelta
    num_of_blocks: int

    def __init__(self, start_time: datetime, end_time: datetime, step_duration: timedelta):
        self.start_time = start_time
        self.end_time = end_time
        self.step_duration = step_duration

        duration_secs = (end_time - start_time).total_seconds()
        step_duration_secs = step_duration.total_seconds()
        if duration_secs % step_duration_secs != 0:
            raise RuntimeError(f'Expected step_duration({step_duration_secs} secs) to fit exactly between '
                               f'start time({start_time}) and end time({end_time})')
        if start_time > end_time:
            raise RuntimeError(f'Cannot create a block metadata where end ({end_time}) is earlier than '
                               f'start ({start_time}).')

        self.num_of_blocks = int(duration_secs / step_duration_secs)

    def convert_to_instant_in_block(self, instant: datetime) -> DecimalInstantInBlock:
        seconds_in_block = (instant - self.start_time).total_seconds()
        factor_in_steps = seconds_in_block / self.step_duration.total_seconds()

        return factor_in_steps

    def convert_to_range_in_block_decimal(self, start: datetime, end: datetime) -> DecimalRangeInBlock:
        if end < start:
            raise RuntimeError(f'Cannot construct range {start}-{end} as start is later then end.')
        # if not self.overlaps(start, end):
        #     raise RuntimeError(f'Will not convert range {start}-{end} as it does not overlap block {self}.')
        return DecimalRangeInBlock(self.convert_to_instant_in_block(start),
                                   self.convert_to_instant_in_block(end))

    def convert_to_range_in_block_int(self, start: datetime, end: datetime) -> IntRangeInBlock:
        decimal_block = self.convert_to_range_in_block_decimal(start, end)
        decimal_start = decimal_block.start
        decimal_end = decimal_block.end

        if decimal_start != int(decimal_start):
            raise RuntimeError(f'Start {start} does not fit exactly on a timestep start on block {self}.')
        if decimal_end != int(decimal_end):
            raise RuntimeError(f'End {end} does not fit exactly on a timestep start on block {self}.')
        return IntRangeInBlock(int(decimal_start), int(decimal_end))

    def overlaps(self, start: datetime, end: datetime) -> bool:
        return times_ranges_overlap(start, end, self.start_time, self.end_time)

    def to_range_in_block_int(self) -> IntRangeInBlock:
        return IntRangeInBlock(0, self.num_of_blocks)

    def from_instant_in_block(self, instant_in_block: DecimalInstantInBlock) -> datetime:
        return self.start_time + (self.step_duration * instant_in_block)

    def from_int_block(self, int_block: IntRangeInBlock) -> Tuple[datetime, datetime]:
        return self.from_instant_in_block(int_block.start), self.from_instant_in_block(int_block.end)


@dataclass
class ValuesInBlockProfile:
    """Abstract profile containing float values per block num"""
    range_in_block: IntRangeInBlock
    value_per_block: list[float]

    def __init__(self, range_in_block: IntRangeInBlock, value_per_block: list[float]):
        self.range_in_block = range_in_block
        self.value_per_block = value_per_block

        if len(value_per_block) != range_in_block.total_block_duration():
            raise RuntimeError(f'Expected the size of the energy_per_block profile ({len(value_per_block)}) and the '
                               f'meta_data ({range_in_block.total_block_duration()}) to be the same!')

    def normalized_index_for_block_num(self, block_num: int) -> int:
        return block_num - self.range_in_block.start

    def contains_value_for_block_num(self, block_num: int) -> bool:
        return self.range_in_block.start <= block_num < self.range_in_block.end

    def normalize_index(self, block_num: int) -> int:
        normalized_index = self.normalized_index_for_block_num(block_num)
        if not self.contains_value_for_block_num(block_num):
            raise RuntimeError(f'Block num {block_num} is outside of profile {self.range_in_block}')
        return normalized_index

    def value_at(self, block_num: int) -> float:
        return self.value_per_block[self.normalize_index(block_num)]


class EnergyProfile(ValuesInBlockProfile):
    """Energy split per block"""
    total_energy: float

    def __init__(self, range_in_block: IntRangeInBlock, energy_per_block: list[float]):
        super().__init__(range_in_block, energy_per_block)
        self.total_energy = sum(energy_per_block)

    def energy_between(self, between: IntRangeInBlock) -> float:
        return sum(self.value_per_block[self.normalize_index(between.start):self.normalized_index_for_block_num(between.end)])

    def energy_at(self, block_num: int) -> float:
        return self.value_at(block_num)

    def mask_int(self, mask: IntRangeInBlock) -> Optional['EnergyProfile']:
        """Generate a new energy profile for the range of steps defined by mask.

        :param mask: The range of steps for which the energy profile should be created.
        :return: An energy profile with the energy values at the steps defined in the mask.
        """
        intersection = self.range_in_block.intersection_int(mask)
        if intersection is not None:
            new_values = []
            for i in intersection.block_nums():
                new_values.append(self.value_at(i))
            return EnergyProfile(intersection, new_values)
        else:
            return None

    def mask_decimal(self, mask: DecimalRangeInBlock) -> Optional['EnergyProfile']:
        """Generate a new energy profile for the range of steps defined by mask.

        Takes the duration for each step into account when determining how much energy at a given step belongs
        to the mask. The duration is used as a factor to multiply the total energy at that step in the profile.

        :param mask:
        :return:
        """
        intersection = self.range_in_block.intersection_int(mask.to_range_in_block_int())
        if intersection is not None:
            new_values = []
            for i in intersection.block_nums():
                duration = mask.duration_at_step_num(i)
                new_values.append(self.value_at(i) * duration)
            return EnergyProfile(intersection, new_values)
        else:
            return None

    def split_on_int_instant(self, instant: IntInstantInBlock) -> tuple[Optional['EnergyProfile'],
                                                                        Optional['EnergyProfile']]:
        left_range, right_range = self.range_in_block.split_on_int_instant(instant)

        left_energy_profile = None
        if left_range:
            left_energy_profile = self.mask_int(left_range)

        right_energy_profile = None
        if right_range:
            right_energy_profile = self.mask_int(right_range)

        return left_energy_profile, right_energy_profile

    def split_on_int(self, split_range: IntRangeInBlock) -> tuple[Optional['EnergyProfile'], Optional['EnergyProfile']]:
        left_range, right_range = self.range_in_block.subtract_int(split_range)

        left_energy_profile = None
        if left_range:
            left_energy_profile = self.mask_int(left_range)

        right_energy_profile = None
        if right_range:
            right_energy_profile = self.mask_int(right_range)

        return left_energy_profile, right_energy_profile

    def profile_concat_right(self, other_right: 'EnergyProfile') -> 'EnergyProfile':
        """ Concat the other profile to the right of this assuming they are aligned.

        The other_right must start immediately on the end of this profile which is what we call aligned.

        :param other_right: The other energy profile to append to the end of this profile.
        :return: A new energy profile where the other is joined to the end of this profile.
        """
        if other_right.range_in_block.start != self.range_in_block.end:
            raise RuntimeError(f'The other energy profile ({other_right}) must start immediately after this '
                               f'profile ({self})')

        new_range_in_block = IntRangeInBlock(self.range_in_block.start, other_right.range_in_block.end)
        new_energy_values = self.value_per_block + other_right.value_per_block

        return EnergyProfile(new_range_in_block, new_energy_values)

    def profile_addition(self, other: 'EnergyProfile') -> 'EnergyProfile':
        common_blocks_start = min(self.range_in_block.start, other.range_in_block.start)
        common_blocks_end = max(self.range_in_block.end, other.range_in_block.end)

        new_values = []
        for i in range(common_blocks_start, common_blocks_end):
            new_value = 0.0

            if self.contains_value_for_block_num(i):
                new_value += self.energy_at(i)

            if other.contains_value_for_block_num(i):
                new_value += other.energy_at(i)

            new_values.append(new_value)

        return EnergyProfile(IntRangeInBlock(common_blocks_start, common_blocks_end), new_values)


class EvFlexMetricProfile(ValuesInBlockProfile):
    """Non-flexible energy split per block"""

    def __init__(self, range_in_block: IntRangeInBlock, flex_metric_per_step: list[float]):
        super().__init__(range_in_block, flex_metric_per_step)

    def flex_metric_at(self, block_num: int) -> float:
        return self.value_at(block_num)


@dataclass
class ChargingSession:
    """The (part of the) charging session that is valid within the block."""
    session: DecimalRangeInBlock
    max_charging_power_watt: float
    max_charging_energy_per_step_joule: float
    energy_to_charge_profile: EnergyProfile
    meta_data: BlockMetadata

    def __init__(self,
                 session: DecimalRangeInBlock,
                 max_charging_power_watt: float,
                 energy_to_charge_profile: EnergyProfile,
                 meta_data: BlockMetadata,
                 fix_energy_profile=False):

        # if session.end > meta_data.num_of_blocks:
        #     raise RuntimeError(f'Charging session end {session.end} is outside of block {meta_data}')
        #
        # if session.start < 0:
        #     raise RuntimeError(f'Charging session start {session.start} is outside of block {meta_data}')
        session_int = session.to_range_in_block_int()
        if energy_to_charge_profile.range_in_block != session_int:
            raise RuntimeError(f'Energy profile {energy_to_charge_profile.range_in_block} runs outside of charging '
                               f'session {session}')

        self.max_charging_energy_per_step_joule = max_charging_power_watt * meta_data.step_duration.total_seconds()
        new_energy_profile = []
        for i in session.block_nums():
            energy_to_charge = energy_to_charge_profile.energy_at(i)
            step_factor = round(session.duration_at_step_num(i), 10)

            max_charging_energy_for_step = step_factor * self.max_charging_energy_per_step_joule
            if energy_to_charge > max_charging_energy_for_step and not fix_energy_profile:
                raise RuntimeError(f'Energy profile charges above max charging capacity at step #{i} with '
                                   f'{energy_to_charge} out of {max_charging_energy_for_step}. '
                                   f'Max joules for charging in a step is {self.max_charging_energy_per_step_joule} '
                                   f'and scoped with factor {step_factor}.')
            new_energy_profile.append(min(max_charging_energy_for_step, energy_to_charge))

        if fix_energy_profile:
            energy_to_charge_profile = EnergyProfile(range_in_block=energy_to_charge_profile.range_in_block,
                                                     energy_per_block=new_energy_profile)

        self.session = session
        self.max_charging_power_watt = max_charging_power_watt
        self.energy_to_charge_profile = energy_to_charge_profile
        self.meta_data = meta_data

    def can_charge_energy_in_step(self, step_num: int) -> float:
        return self.max_charging_energy_per_step_joule * self.session.duration_at_step_num(step_num)

    def non_flexible_energy_utilizing_whole_session(self, congestion_steps: IntRangeInBlock) -> Optional[float]:
        """Calculate the amount of joules that cannot be charged in non-congested blocks utilizing
            the whole session duration.

        :param congestion_steps: Steps within block which have congestion.
        :return: The amount of joules that cannot be charged in non-congested blocks.
        """
        session_congestion = self.session.intersection_decimal(congestion_steps)

        if not session_congestion:
            non_flexible_energy = None
        else:
            non_congestion_room_joule = (self.session.total_block_duration() - session_congestion.total_block_duration()) * self.max_charging_energy_per_step_joule

            non_flexible_energy = max(self.energy_to_charge_profile.total_energy - non_congestion_room_joule, 0)

        return non_flexible_energy

    def non_flexible_energy_utilizing_after_congestion(self,
                                                       flex_window: BlockMetadata,
                                                       congestion_steps: IntRangeInBlock) -> Optional[float]:
        """Calculate the amount of joules that cannot be charged in non-congested blocks utilizing
            only the session duration after congestion.

        :param congestion_steps: Steps within block which have congestion.
        :return: The amount of joules that cannot be charged in non-congested blocks.
        """
        session_during_flex_window = self.session.intersection_decimal(flex_window.to_range_in_block_int())
        if not session_during_flex_window:
            non_flexible_energy = None
        else:
            session_during_congestion = session_during_flex_window.intersection_decimal(congestion_steps)
            _, session_after_congestion = session_during_flex_window.subtract_decimal(congestion_steps)

            if not session_during_congestion:
                non_flexible_energy = None
            elif not session_after_congestion:
                non_flexible_energy = self.energy_to_charge_profile.energy_between(session_during_congestion.to_range_in_block_int())
            else:
                charged_energy_in_congestion = self.energy_to_charge_profile.energy_between(session_during_congestion.to_range_in_block_int())
                default_charged_energy_after_congestion = self.energy_to_charge_profile.energy_between(session_after_congestion.to_range_in_block_int())

                non_congestion_room_joule = max((session_after_congestion.total_block_duration() * self.max_charging_energy_per_step_joule) - default_charged_energy_after_congestion,
                                                0)
                non_flexible_energy = max(charged_energy_in_congestion - non_congestion_room_joule, 0)

        return non_flexible_energy

    def non_flexible_energy_evenly_divided(self,
                                           non_flexible_energy: float,
                                           congestion_steps: IntRangeInBlock) -> Optional[EnergyProfile]:
        """Evenly divide the non_flexible_energy across the steps in congestion_steps.

        :param non_flexible_energy: The amount of energy which is considered non flexible and should remain in the
                                    congestion steps.
        :param congestion_steps: Steps within block which have congestion.
        :return: The new energy profile where the non_flexible_energy is divided across congestion_steps.
        """
        session_congestion_dec = self.session.intersection_decimal(congestion_steps)

        if session_congestion_dec is not None and non_flexible_energy is not None:
            non_flexible_energy_per_block = non_flexible_energy / session_congestion_dec.total_block_duration()

            energy_per_block = []
            for i in session_congestion_dec.block_nums():
                step_duration = session_congestion_dec.duration_at_step_num(i)
                energy_per_block.append(non_flexible_energy_per_block * step_duration)

            return EnergyProfile(session_congestion_dec.to_range_in_block_int(),
                                 energy_per_block)
        else:
            return None

    def non_flexible_energy_evenly_divided_while_not_increasing_above_default_charging(self,
                                                                                       non_flexible_energy: float,
                                                                                       congestion_steps: IntRangeInBlock) -> Optional[EnergyProfile]:
        """Evenly distribute the non-flexible energy across the congestion steps while not increasing
            above the default charging on each step.

        :param non_flexible_energy: The amount of energy which is considered non flexible and should remain in the
            congestion steps.
        :param congestion_steps: Steps within block which have congestion.
        :return: The energy profile during the congestion where the non flexible energy charged is evenly divided
                 across the steps but each step will not charge more than the default charging profile.
        """
        session_congestion_dec = self.session.intersection_decimal(congestion_steps)

        if session_congestion_dec is not None and non_flexible_energy is not None:
            # First sort blocks by least default_energy_first
            least_default_energy_first: list[tuple[int, float, float]] = []
            for i in session_congestion_dec.block_nums():
                default_energy = self.energy_to_charge_profile.energy_at(i)
                step_duration = session_congestion_dec.duration_at_step_num(i)
                least_default_energy_first.append((i, default_energy, step_duration))
            least_default_energy_first.sort(key=lambda i: i[1])

            non_flexible_energy_to_divide = non_flexible_energy
            total_remaining_duration = len(least_default_energy_first)
            energy_per_block = []
            for step_num, default_energy_at_step, step_duration in least_default_energy_first:
                proposed_non_flexible_energy_current_block = non_flexible_energy_to_divide / total_remaining_duration
                set_non_flexible_energy_current_block = min(default_energy_at_step,
                                                            proposed_non_flexible_energy_current_block)
                energy_per_block.append((step_num, set_non_flexible_energy_current_block))
                non_flexible_energy_to_divide = non_flexible_energy_to_divide - set_non_flexible_energy_current_block
                total_remaining_duration -= 1
            energy_per_block.sort(key=lambda i: i[0])

            result = EnergyProfile(session_congestion_dec.to_range_in_block_int(),
                                 list(map(lambda i: i[1], energy_per_block)))

            return result
        else:
            return None

    def charge_extra_energy_immediately(self,
                                        energy_profile: EnergyProfile,
                                        energy_joule: float) -> EnergyProfile:
        energy_to_charge = energy_joule
        new_energy_values = []
        for i in energy_profile.range_in_block.block_nums():
            energy_in_step = energy_profile.energy_at(i)
            energy_room = self.can_charge_energy_in_step(i) - energy_in_step
            will_charge_extra = min(energy_room, energy_to_charge)
            energy_to_charge -= will_charge_extra
            new_energy_values.append(energy_in_step + will_charge_extra)

        if energy_to_charge > 0.001:
            raise RuntimeError(f'Could not fit {energy_to_charge} out of {energy_joule} in energy profile {self}')

        return EnergyProfile(energy_profile.range_in_block,
                             new_energy_values)

    def shift_flexible_energy_after_congestion(self,
                                               flex_window: BlockMetadata,
                                               congestion_steps: IntRangeInBlock) -> EnergyProfile:
        """Shift as much flexible energy outside of the congestion to after the congestion steps.

        Energy is shifted to immediately after the congestion and charging as quickly as possible.

        :param congestion_steps: Steps within block which have congestion.
        :return: The alternative energy profile where the energy during congestion is shifted to immediately
            after the congestion.
        """
        non_flexible_energy = self.non_flexible_energy_utilizing_after_congestion(flex_window,
                                                                                  congestion_steps)
        congestion_energy_profile = self.non_flexible_energy_evenly_divided_while_not_increasing_above_default_charging(non_flexible_energy,
                                                                                                                        congestion_steps)
        if non_flexible_energy is None and congestion_energy_profile is None:
            return self.energy_to_charge_profile

        if not math.isclose(non_flexible_energy, congestion_energy_profile.total_energy, rel_tol=0.01):
            raise RuntimeError(f'The non flexible energy {non_flexible_energy} does not match the desired energy '
                               f'profile {congestion_energy_profile}(total_energy: '
                               f'{congestion_energy_profile.total_energy}) during congestion. This should not happen.')

        default_profile_during_congestion = self.energy_to_charge_profile.mask_int(congestion_steps)
        energy_to_move = default_profile_during_congestion.total_energy - congestion_energy_profile.total_energy

        default_profile_before_congestion, default_profile_after_congestion = self.energy_to_charge_profile.split_on_int(congestion_steps)

        if default_profile_before_congestion:
            resulting_energy_profile = default_profile_before_congestion.profile_concat_right(congestion_energy_profile)
        else:
            resulting_energy_profile = congestion_energy_profile
            #default_profile_before_block, default_profile_during_block_before_congestion = default_profile_before_congestion.split_on_int(self.meta_data.to_range_in_block_int())
        if default_profile_after_congestion:
            default_profile_during_block_after_congestion, default_profile_after_block = default_profile_after_congestion.split_on_int_instant(flex_window.to_range_in_block_int().end)
            shifted_profile_after_congestion = self.charge_extra_energy_immediately(default_profile_during_block_after_congestion,
                                                                                    energy_to_move)
            resulting_energy_profile = resulting_energy_profile.profile_concat_right(shifted_profile_after_congestion)
            if default_profile_after_block:
                resulting_energy_profile = resulting_energy_profile.profile_concat_right(default_profile_after_block)
        elif energy_to_move > (0.01 * default_profile_during_congestion.total_energy):
            raise RuntimeError(f'There was energy to move ({energy_to_move}) after the congestion but there is no '
                               f'profile after the congestion')

        return resulting_energy_profile

    def calculate_flex_metric(self, congestion_steps: IntRangeInBlock) -> EvFlexMetricProfile:
        ''' Calculates the ev_flex_metric for a single EV session.
        DEPRECATED

        :param congestion_steps:
        :return:
        '''
        non_flexible_energy = self.non_flexible_energy_utilizing_whole_session(congestion_steps)
        flex_steps = self.energy_to_charge_profile.range_in_block.intersection_int(congestion_steps)

        ev_flex_metric_per_step = []
        if non_flexible_energy is not None:

            original_energy = 0
            for i in flex_steps.block_nums():
                original_energy_step = self.energy_to_charge_profile.energy_at(i)
                original_energy += original_energy_step

            if original_energy:
                ev_flex_metric = (original_energy - non_flexible_energy) / original_energy
            else:
                ev_flex_metric = 0
            ev_flex_metric_per_step = [ev_flex_metric] * flex_steps.total_block_duration()

                # non_flexible_energy_per_step = non_flexible_energy / flex_steps.total_block_duration() # TODO mistake: Should use block duration of decimal, niet int range in block.
                # if original_energy_step:
                #     ev_flex_metric_step = (original_energy_step - non_flexible_energy_per_step) / original_energy_step
                # else:
                #     ev_flex_metric_step = 0
                # ev_flex_metric_per_step.append(ev_flex_metric_step)

        return EvFlexMetricProfile(flex_steps, ev_flex_metric_per_step)


def to_energy_profile_using_default_charge_behaviour(session: DecimalRangeInBlock,
                                                     block_metadata: BlockMetadata,
                                                     charged_energy_kwh: float,
                                                     charging_time: timedelta,
                                                     max_power_kw: float) -> EnergyProfile:
    blocks_charging = charging_time.total_seconds() / block_metadata.step_duration.total_seconds()
    energy_per_block_joule = charged_energy_kwh * 3_600_000 / blocks_charging
    max_power_watt = max_power_kw * 1000

    max_energy_per_timestep = max_power_watt * block_metadata.step_duration.total_seconds()

    one_watt_for_timestep_joule = block_metadata.step_duration.total_seconds()

    if (energy_per_block_joule - max_energy_per_timestep) > one_watt_for_timestep_joule:
        raise RuntimeError(f'Discrepancy between max charge power ({max_power_watt} watt) and average '
                           f'energy charged per timestep '
                           f'({energy_per_block_joule / block_metadata.step_duration.total_seconds()} watt)')
    elif energy_per_block_joule > max_energy_per_timestep:
        energy_per_block_joule = max_energy_per_timestep

    energy_per_block = []

    blocks_left_to_charge = blocks_charging
    start_partial_factor = round(min(blocks_left_to_charge,
                                     min(math.ceil(session.start),
                                         session.end) - session.start),
                                 10)
    if start_partial_factor:
        energy_per_block.append(start_partial_factor * energy_per_block_joule)
        blocks_left_to_charge = round(blocks_left_to_charge - start_partial_factor, 10)

    blocks_charging_fully = max(0, math.floor(blocks_left_to_charge))
    energy_per_block.extend([energy_per_block_joule] * blocks_charging_fully)

    last_partial_block_factor = blocks_left_to_charge % 1
    if last_partial_block_factor:
        energy_per_block.append(last_partial_block_factor * energy_per_block_joule)

    session_int = session.to_range_in_block_int()

    return EnergyProfile(session_int,
                         energy_per_block + [0] * (session_int.total_block_duration() - len(energy_per_block)))


@dataclass
class ElaadChargingSession:
    """
    ",  "TransactionId","ChargePoint","Connector","UTCTransactionStart","UTCTransactionStop","StartCard",                                                       "ConnectedTime","ChargeTime","IdleTime","TotalEnergy","MaxPower"
    "1","3327068",      "AL111",      "AL111-1",  2019-03-01 11:50:37,  2019-03-01 13:21:44, "6cfef3fda701fb605ea8f4cdd9d4700b2778683d846655fab91d797c58981b42",1.52,           1.52,        0,         6.81,          4.909

    """
    ALLOWED_TRANSACTION_DURATION_INCREASE_TO_FIT_CHARGING_FACTOR = 0.01

    transaction_id: str
    utc_session_start: datetime
    utc_session_stop: datetime

    charging_time: timedelta
    charged_energy_kwh: float
    max_power_kw: float

    def to_general_charging_session(self, block_metadata: BlockMetadata) -> Optional[ChargingSession]:
        block_range = block_metadata.to_range_in_block_int()
        session = block_metadata.convert_to_range_in_block_decimal(self.utc_session_start, self.utc_session_stop)
        energy_profile = to_energy_profile_using_default_charge_behaviour(session,
                                                                          block_metadata,
                                                                          self.charged_energy_kwh,
                                                                          self.charging_time,
                                                                          self.max_power_kw)

        if session:
            masked_session = session.intersection_decimal(block_range)
            assert masked_session is not None
            masked_energy_profile = energy_profile.mask_int(block_range)
            assert masked_energy_profile is not None
            return ChargingSession(masked_session,
                                   self.max_power_kw * 1000,
                                   masked_energy_profile,
                                   block_metadata)
        else:
            return None

    @staticmethod
    def from_line(line: str) -> Optional['ElaadChargingSession']:
        parts = line.rstrip('\n').split(',')
        skip_reasons = []

        if parts[8] == 'NA':
            skip_reasons.append('Charge duration was NA instead of a number.')

        if parts[10] == 'NA':
            skip_reasons.append('Charged energy in kWh was NA instead of a number.')

        if parts[11] == 'NA':
            skip_reasons.append('Peak power in kW was NA instead of a number.')

        if parts[8] == '0' and parts[10] != '0':
            skip_reasons.append(f'Charge time was 0 but there was {parts[10]} energy charged.')

        if skip_reasons:
            print(f'Skipping unparseable line (reasons: {skip_reasons}): {line}')
            result = None
        else:
            transaction_id = parts[1].strip('"')
            start = datetime.fromisoformat(parts[4]).replace(tzinfo=pytz.utc)
            end = datetime.fromisoformat(parts[5]).replace(tzinfo=pytz.utc)

            transaction_duration = end - start
            charging_time = timedelta(hours=float(parts[8]))

            if 0 < (charging_time.total_seconds() - transaction_duration.total_seconds()) <= 36:
                # Charging time exceeds transaction duration.
                # Margin of 36 seconds is acceptable as the charge time is in hours and only significant to 2 digits
                # behind the comma. This is a resolution of 36 seconds.
                print(f'Warning! [{transaction_id}] Transaction duration ({transaction_duration.total_seconds()} '
                      f'seconds) was less than charging time ({charging_time.total_seconds()} seconds) but within '
                      f'acceptable margin. Correct transaction duration to charging time.')
                end = start + charging_time
            elif (charging_time.total_seconds() - transaction_duration.total_seconds()) >= 36:
                raise RuntimeError(f'[{transaction_id}] charging time {charging_time} was significantly longer than '
                                   f'transaction duration {transaction_duration}. Dataset has a bad transaction.')

            charged_energy_kwh = float(parts[10])
            max_power_kw = float(parts[11])

            average_power_kw = charged_energy_kwh / (charging_time.total_seconds() / 3600)

            if average_power_kw > max_power_kw:
                # Charging time was not long enough to stay below max_power_kw.
                new_charging_time = timedelta(hours=charged_energy_kwh / max_power_kw)
                increased_charging_time = new_charging_time.total_seconds() - charging_time.total_seconds()
                non_charge_seconds_in_transaction = (transaction_duration - charging_time).total_seconds()
                max_increase_allowed = max(36, non_charge_seconds_in_transaction + ElaadChargingSession.ALLOWED_TRANSACTION_DURATION_INCREASE_TO_FIT_CHARGING_FACTOR * transaction_duration.total_seconds())

                if 0 <= increased_charging_time <= max_increase_allowed:
                    print(f'Warning! [{transaction_id}] Charging duration was not long enough causing the charging '
                          f'session to use more than max_power_kw on average. Increasing charging_time by '
                          f'{math.ceil(increased_charging_time)} seconds which is within acceptable margin.')
                    charging_time = charging_time + timedelta(seconds=math.ceil(increased_charging_time))
                    end = max(end, start + charging_time)
                elif increased_charging_time >= max_increase_allowed:
                    raise RuntimeError(f'[{transaction_id}] could not increase charge time enough to lower '
                                       f'the average charging power ({average_power_kw} kwatt) below the max power '
                                       f'({max_power_kw} kwatt). Wanted to increase charging '
                                       f'time with {increased_charging_time} seconds but was only allowed '
                                       f'{max_increase_allowed} seconds. Dataset has a bad transaction.')
                else:
                    raise RuntimeError('Something weird happened...')
    
            result = ElaadChargingSession(transaction_id, start, end, charging_time, charged_energy_kwh, max_power_kw)

        return result

    @staticmethod
    def parse_file(path: Path) -> list['ElaadChargingSession']:
        elaad_transactions = []
        num_skipped = 0
        total_lines = 0
        with open(path) as open_file:
            open_file.readline() # skip header
            for line in open_file:
                transaction = ElaadChargingSession.from_line(line)
                if transaction is None:
                    num_skipped += 1
                else:
                    elaad_transactions.append(transaction)
                total_lines += 1

        print(f'Skipped {num_skipped} out of {total_lines} lines.')

        return elaad_transactions


@dataclass
class AlbatrosChargingSession:
    """
    ",  "TransactionId","ChargePoint","Connector","UTCTransactionStart","UTCTransactionStop","StartCard",                                                       "ConnectedTime","ChargeTime","IdleTime","TotalEnergy","MaxPower"
    "1","3327068",      "AL111",      "AL111-1",  2019-03-01 11:50:37,  2019-03-01 13:21:44, "6cfef3fda701fb605ea8f4cdd9d4700b2778683d846655fab91d797c58981b42",1.52,           1.52,        0,         6.81,          4.909

    """
    ALLOWED_TRANSACTION_DURATION_INCREASE_TO_FIT_CHARGING_FACTOR = 0.01

    transaction_id: str
    utc_session_start: datetime
    utc_session_stop: datetime

    charging_time: timedelta
    charged_energy_kwh: float
    max_power_kw: float

    def to_general_charging_session(self, block_metadata: BlockMetadata) -> Optional[ChargingSession]:
        block_range = block_metadata.to_range_in_block_int()
        session = block_metadata.convert_to_range_in_block_decimal(self.utc_session_start, self.utc_session_stop)
        energy_profile = to_energy_profile_using_default_charge_behaviour(session,
                                                                          block_metadata,
                                                                          self.charged_energy_kwh,
                                                                          self.charging_time,
                                                                          self.max_power_kw)

        if session:
            masked_session = session.intersection_decimal(block_range)
            assert masked_session is not None
            masked_energy_profile = energy_profile.mask_int(block_range)
            assert masked_energy_profile is not None
            return ChargingSession(masked_session,
                                   self.max_power_kw * 1000,
                                   masked_energy_profile,
                                   block_metadata)
        else:
            return None

    @staticmethod
    def from_line(line: dict) -> Optional['AlbatrosChargingSession']:
        transaction_id = int(line['session_id'])
        start = datetime.fromisoformat(line['startTime']).replace(tzinfo=pytz.utc)
        end_charge_time = datetime.fromisoformat(line['endTime']).replace(tzinfo=pytz.utc)
        end = datetime.fromisoformat(line['plugOutTime']).replace(tzinfo=pytz.utc)

        transaction_duration = end - start
        charging_time = end_charge_time - start

        if charging_time.total_seconds() > transaction_duration.total_seconds():
            raise RuntimeError(f'[{transaction_id}] charging time {charging_time} was longer than '
                               f'transaction duration {transaction_duration}. Dataset has a bad transaction.')

        charged_energy_kwh = line['charge_kWh']
        max_power_kw = line['maxChargePower_kW']

        average_power_kw = charged_energy_kwh / (charging_time.total_seconds() / 3600)

        if (average_power_kw - max_power_kw) > 0.001:
            raise RuntimeError(f'[{transaction_id}] the average charging power ({average_power_kw} kwatt) below '
                               f'the max power ({max_power_kw} kwatt). Dataset has a bad transaction.')

        result = AlbatrosChargingSession(transaction_id, start, end, charging_time, charged_energy_kwh, max_power_kw)

        return result

    @staticmethod
    def parse_file(path: Path) -> list['AlbatrosChargingSession']:
        transactions = []
        num_skipped = 0
        total_lines = 0

        workbook = openpyxl.load_workbook(path, keep_vba=False, data_only=True)
        sheet = workbook.active

        headers = []
        for column in sheet.iter_cols(min_row=1, max_row=2):
            headers.append(column[0].value)
        lines = []
        for row in sheet.iter_rows(min_row=2):
            line = {}
            for header, cell in zip(headers, row):
                line[header] = cell.value
            lines.append(line)

        for line in lines:
            transaction = AlbatrosChargingSession.from_line(line)
            if transaction is None:
                num_skipped += 1
            else:
                transactions.append(transaction)
            total_lines += 1

        print(f'Skipped {num_skipped} out of {total_lines} lines.')

        return transactions


def calculate_ev_flex_metric(block_metadata: BlockMetadata,
                             congestion: IntRangeInBlock,
                             sessions_in_block: list[ChargingSession]) -> Optional[EvFlexMetricProfile]:
    if not block_metadata.to_range_in_block_int().contains(congestion):
        raise RuntimeError(f'Congestion range {congestion} should be contained by block {block_metadata}.')

    total_non_flexible_energy_profile: Optional[EnergyProfile] = None
    total_default_energy_profile: Optional[EnergyProfile] = None
    for charging_session in sessions_in_block:
        non_flexible_energy = charging_session.non_flexible_energy_utilizing_whole_session(congestion)
        non_flexible_energy_evenly_divided = charging_session.non_flexible_energy_evenly_divided(non_flexible_energy,
                                                                                                 congestion)

        if total_non_flexible_energy_profile is None:
            total_non_flexible_energy_profile = non_flexible_energy_evenly_divided
        elif non_flexible_energy_evenly_divided is not None:
            total_non_flexible_energy_profile = total_non_flexible_energy_profile.profile_addition(non_flexible_energy_evenly_divided)

        default_energy_profile = charging_session.energy_to_charge_profile.mask_int(congestion)
        if total_default_energy_profile is None:
            total_default_energy_profile = default_energy_profile
        elif default_energy_profile is not None:
            total_default_energy_profile = total_default_energy_profile.profile_addition(default_energy_profile)

    if total_default_energy_profile is not None and total_non_flexible_energy_profile is not None:
        if total_non_flexible_energy_profile.range_in_block != total_default_energy_profile.range_in_block:
            raise RuntimeError('Something weird happened...')

        ev_flex_values = []
        for i in total_default_energy_profile.range_in_block.block_nums():
            non_flex_energy = total_non_flexible_energy_profile.energy_at(i)
            default_energy = total_default_energy_profile.energy_at(i)
            if default_energy == 0.0:
                ev_flex_metric = 0
            else:
                ev_flex_metric = (default_energy - non_flex_energy) / default_energy
            ev_flex_values.append(ev_flex_metric)

        return EvFlexMetricProfile(total_default_energy_profile.range_in_block,
                                   ev_flex_values)
    return None


def main():
    # transactions = ElaadChargingSession.parse_file(Path('/mnt/vm-shared/ElaadNL datasets.HoogVertrouwelijk/transactions1Y.csv'))
    transactions = AlbatrosChargingSession.parse_file(Path('/mnt/vm-shared/ChargeSessions_private_charging_5501.xlsx'))
    output_schema = json.dumps({'name': 'ev_flex_metric_elaad_2019',
                                'type': 'record',
                                'fields': [{"name": "block_start_epoch_timestamp",
                                            "type": {'type': 'long',
                                                     "logicalType": "timestamp-millis"}},
                                           {"name": "timestep_duration_seconds",
                                            "type": "int"},
                                           {"name": "block_length_timestep",
                                            "type": "int"},
                                           {"name": "congestion_start_timestep",
                                            "type": "int"},
                                           {"name": "congestion_end_timestep",
                                            "type": "int"},
                                           {"name": "ev_flex_metric_for_timestep",
                                            "type": 'int'},
                                           {"name": "ev_flex_metric_value",
                                            "type": 'double'},
                                           {"name": "num_of_charging_sessions_during_congestion",
                                            "type": "int"},
                                           ]})

    with DataFileWriter(open('notebooks/ev_flex_metric.avro', "wb"), DatumWriter(), avro.schema.parse(output_schema)) as writer:
        resolution = timedelta(hours=1)
        step_duration = timedelta(minutes=15)
        for block_length_duration in [8, 12, 16, 20, 24, 28, 32, 36]:
            for congestion_start in range(0, block_length_duration, 4):
                start = datetime(year=2021, month=6, day=1, hour=0, minute=0, second=0, tzinfo=pytz.utc)
                final = start + timedelta(days=7)
                block_duration = step_duration * block_length_duration
                end = start + block_duration
                block_metadata = BlockMetadata(start, end, step_duration)
                congestion = IntRangeInBlock(congestion_start, congestion_start + 4)

                while start < final:
                    congestion_start = congestion.start * step_duration + start
                    congestion_end = congestion.end * step_duration + start
                    charging_sessions = []
                    for elaad_transaction in transactions:
                        if times_ranges_overlap(congestion_start,
                                                congestion_end,
                                                elaad_transaction.utc_session_start,
                                                elaad_transaction.utc_session_stop):
                            charging_session = elaad_transaction.to_general_charging_session(block_metadata)
                            if charging_session:
                                charging_sessions.append(charging_session)
                    print(f'A number of {len(charging_sessions)} charging sessions overlap with the chosen block')

                    ev_flex_metric = calculate_ev_flex_metric(block_metadata, congestion, charging_sessions)
                    print(f'{start}-{end}', ev_flex_metric)
                    if ev_flex_metric:
                        for ev_flex_metric_timestep, ev_flex_matric_value in zip(range(congestion.start, congestion.end),
                                                                                 ev_flex_metric.value_per_block):
                            writer.append({'block_start_epoch_timestamp': start,
                                           'timestep_duration_seconds': int(step_duration.total_seconds()),
                                           'block_length_timestep': block_metadata.num_of_blocks,
                                           'congestion_start_timestep': congestion.start,
                                           'congestion_end_timestep': congestion.end,
                                           'ev_flex_metric_for_timestep': ev_flex_metric_timestep,
                                           'ev_flex_metric_value': ev_flex_matric_value,
                                           'num_of_charging_sessions_during_congestion': len(charging_sessions)})

                    start = start + resolution
                    end = start + block_duration
                    block_metadata = BlockMetadata(start, end, step_duration)

                    # congestion = IntRangeInBlock(2, 4)
                    # print(calculate_ev_flex_metric(block_metadata, congestion, charging_sessions))
                    #
                    # congestion = IntRangeInBlock(3, 4)
                    # print(calculate_ev_flex_metric(block_metadata, congestion, charging_sessions))
                    #
                    # congestion = IntRangeInBlock(0, 4)
                    # print(calculate_ev_flex_metric(block_metadata, congestion, charging_sessions))


if __name__ == '__main__':
    main()
